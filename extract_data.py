import openpyxl
import json
import re
from datetime import datetime, time, timedelta

XLSX = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx'


def serialize(v):
    if isinstance(v, datetime):
        return v.isoformat()
    elif isinstance(v, time):
        return v.strftime('%H:%M:%S')
    elif isinstance(v, timedelta):
        total = int(v.total_seconds())
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return f"{h:02d}:{m:02d}:{s:02d}"
    elif isinstance(v, str) and v in ('#DIV/0!', '#REF!', 'LICENÇA', 'LICENÇA'):
        return None
    return v

MONTH_NORM_MAP = {
    'JANEIRO': 'janeiro', 'FEVEREIRO': 'fevereiro',
    'MARCO': 'marco', 'MARÇO': 'marco',
    'ABRIL': 'abril', 'MAIO': 'maio', 'JUNHO': 'junho',
    'JULHO': 'julho', 'AGOSTO': 'agosto', 'SETEMBRO': 'setembro',
    'OUTUBRO': 'outubro', 'NOVEMBRO': 'novembro', 'DEZEMBRO': 'dezembro',
}

def strip_acc(s):
    """Remove common Portuguese accents for comparison."""
    return (s.replace('Ç','C').replace('ç','c')
             .replace('Ã','A').replace('ã','a')
             .replace('Á','A').replace('á','a')
             .replace('Â','A').replace('â','a')
             .replace('É','E').replace('é','e')
             .replace('Ê','E').replace('ê','e')
             .replace('Í','I').replace('í','i')
             .replace('Ó','O').replace('ó','o')
             .replace('Ô','O').replace('ô','o')
             .replace('Ú','U').replace('ú','u'))

def normalize_mes(v):
    """Normalize a month value (string or datetime) to accent-free lowercase, e.g. 'MARÇO' → 'marco'."""
    if v is None:
        return None
    
    months_list = ['janeiro', 'fevereiro', 'marco', 'abril', 'maio', 'junho', 
                   'julho', 'agosto', 'setembro', 'outubro', 'novembro', 'dezembro']
    
    if isinstance(v, datetime):
        return months_list[v.month - 1]
    
    s = str(v).strip()
    key = strip_acc(s).upper()
    
    # Check for direct matches in MONTH_NORM_MAP
    if key in MONTH_NORM_MAP:
        return MONTH_NORM_MAP[key]
        
    # Try prefix match (e.g. 'MAR' → 'marco')
    for k, norm in MONTH_NORM_MAP.items():
        if key.startswith(k[:3]):
            return norm
            
    return strip_acc(s).lower()

wb = openpyxl.load_workbook(XLSX, data_only=True)

all_data = {}

# =============================================================
# ABS DATA - REMOVED AS REQUESTED
# =============================================================
all_data['abs_data'] = {}

# =============================================================
# PRODUÇÃO OPERADORES (individual operators)
# =============================================================
ws = wb['PRODUÇÃO OPERADORES']
operators = []

# Map headers dynamically to avoid breakage if columns shift
header_row = [str(c).upper().strip() if c is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
idx_agente = header_row.index('AGENTE') if 'AGENTE' in header_row else 0
idx_op = header_row.index('OPERAÇÃO') if 'OPERAÇÃO' in header_row else 1
idx_matricula = header_row.index('MATRICULA') if 'MATRICULA' in header_row else 4
idx_ho = header_row.index('H.O') if 'H.O' in header_row else 5
idx_alc_ho = header_row.index('ALCANCE H.O') if 'ALCANCE H.O' in header_row else 6
idx_disp = header_row.index('DISPERSÃO') if 'DISPERSÃO' in header_row else 7
idx_prom = header_row.index('PROMESSA') if 'PROMESSA' in header_row else 10
idx_qual = header_row.index('QUALIDADE') if 'QUALIDADE' in header_row else 13
idx_abs = header_row.index('ABS') if 'ABS' in header_row else 15
idx_q_ho = header_row.index('QUARTIL H.O') if 'QUARTIL H.O' in header_row else 16
idx_tl = header_row.index('TEMPO LOGADO') if 'TEMPO LOGADO' in header_row else 18
idx_pausa = header_row.index('PAUSA 100%') if 'PAUSA 100%' in header_row else 22
idx_mes = header_row.index('MÊS') if 'MÊS' in header_row else 23

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[idx_agente] is None:
        continue
        
    mes_val = row[idx_mes] if idx_mes < len(row) else None
    row_year = None
    if isinstance(mes_val, datetime):
        row_year = mes_val.year
    elif isinstance(mes_val, str):
        m = re.search(r'202\d', mes_val)
        if m: row_year = int(m.group(0))
    
    # Clean Quartil: "1º" -> "1"
    q_raw = str(row[idx_q_ho]) if idx_q_ho < len(row) and row[idx_q_ho] is not None else ""
    q_match = re.search(r'\d', q_raw)
    q_val = q_match.group(0) if q_match else None

    rec = {
        'agente':       str(row[idx_agente]).strip(),
        'ho':           row[idx_ho] if idx_ho < len(row) else 0,
        'alcance_ho':   row[idx_alc_ho] if idx_alc_ho < len(row) else 0,
        'dispersao':    row[idx_disp] if idx_disp < len(row) else 0,
        'promessa':     row[idx_prom] if idx_prom < len(row) else 0,
        'qualidade':    row[idx_qual] if idx_qual < len(row) else 0,
        'abs':          row[idx_abs] if idx_abs < len(row) else 0,
        'quartil_ho':   q_val,
        'tempo_logado': serialize(row[idx_tl]) if idx_tl < len(row) else None,
        'pausa_100':    row[idx_pausa] if idx_pausa < len(row) else 0,
        'mes':          normalize_mes(mes_val),
        'ano':          row_year,
        'operacao':     str(row[idx_op]).strip() if idx_op < len(row) and row[idx_op] else None,
        'matricula':    row[idx_matricula] if idx_matricula < len(row) else None
    }
    
    if rec.get('mes'):
        operators.append(rec)
all_data['operadores'] = operators

# =============================================================
# PRODUÇÃO OPERAÇÕES (team-level KPIs per carteira+operação+mês)
# Col: CARTEIRA | Operação | H.O | DISPERSÃO | ALÔ | CPC | PROMESSA | QUALIDADE | ABS | Tempo logado | % tempo log | Pausa | %pausa | MÊS
# =============================================================
ws = wb['PRODUÇÃO OPERAÇÕES']
prod_ops = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None or row[1] is None:
        continue
    # Ignorar linha de cabeçalho repetida
    if str(row[0]).upper() == 'CARTEIRA':
        continue
    # Detect year from row[13] (MÊS)
    row_year = None
    if isinstance(row[13], datetime):
        row_year = row[13].year
    elif isinstance(row[13], str):
        m = re.search(r'202\d', row[13])
        if m: row_year = int(m.group(0))

    rec = {
        'carteira':     str(row[0]).strip() if row[0] else None,
        'operacao':     str(row[1]).strip() if row[1] else None,
        'ho':           row[2] if isinstance(row[2], (int, float)) else None,
        'dispersao':    serialize(row[3]),
        'alo':          row[4] if isinstance(row[4], (int, float)) else None,
        'cpc':          row[5] if isinstance(row[5], (int, float)) else None,
        'promessa':     row[6] if isinstance(row[6], (int, float)) else None,
        'qualidade':    serialize(row[7]),
        'abs':          serialize(row[8]),
        'tempo_logado': serialize(row[9]),
        'pct_tempo_log':serialize(row[10]),
        'pausa':        serialize(row[11]),
        'pct_pausa':    serialize(row[12]),
        'mes':          normalize_mes(row[13]),
        'ano':          row_year
    }
    if rec.get('mes') and rec.get('operacao'):
        prod_ops.append(rec)
all_data['producao_operacoes'] = prod_ops

# =============================================================
# FECHAMENTOS 2026
# Col: MÊS | OPERAÇÃO | HONORARIOS (H.O) | META | ALCANCE | PA | META PA | ...
# =============================================================
ws = wb['FECHAMENTOS 2026']
fechamentos2026 = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None or row[1] is None:
        continue
    mes_val = normalize_mes(row[0])
    if not mes_val:
        continue
    rec = {
        'mes':      mes_val,
        'operacao': str(row[1]).strip() if row[1] else None,
        'ho':       row[2] if isinstance(row[2], (int, float)) else None,  # HONORARIOS
        'meta':     row[3] if isinstance(row[3], (int, float)) else None,
        'alcance':  row[4] if isinstance(row[4], (int, float)) else None,
        'pa':       row[5] if isinstance(row[5], (int, float)) else None,
        'meta_pa':  row[6] if isinstance(row[6], (int, float)) else None,
        'ano':      2026
    }
    if rec.get('operacao'):
        fechamentos2026.append(rec)
all_data['fechamentos2026'] = fechamentos2026

# =============================================================
# META 2025
# Col: MÊS | OPERAÇÃO | HONORARIOS | META | ALCANCE | PA | META PA | ...
# =============================================================
ws = wb['META 2025']
meta2025 = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None:
        continue
    rec = {
        'mes':        normalize_mes(row[0]),
        'operacao':   str(row[1]).strip() if row[1] else None,
        'ho':         row[2] if isinstance(row[2], (int, float)) else None,
        'meta':       row[3] if isinstance(row[3], (int, float)) else None,
        'alcance':    row[4] if isinstance(row[4], (int, float)) else None,
        'pa':         row[5] if isinstance(row[5], (int, float)) else None,
        'meta_pa':    row[6] if isinstance(row[6], (int, float)) else None,
        'ano':        2025
    }
    if rec.get('operacao'):
        meta2025.append(rec)
all_data['meta2025'] = meta2025

# =============================================================
# META 2024
# Col: MÊS | OPERAÇÃO | HONORARIOS | META | PROJEÇÃO | ALCANCE
# =============================================================
ws = wb['META 2024']
meta2024 = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None:
        continue
    rec = {
        'mes':        normalize_mes(row[0]),
        'operacao':   str(row[1]).strip() if row[1] else None,
        'ho':         row[2] if isinstance(row[2], (int, float)) else None,
        'meta':       row[3] if isinstance(row[3], (int, float)) else None,
        'projecao':   row[4] if isinstance(row[4], (int, float)) else None,
        'alcance':    row[5] if isinstance(row[5], (int, float)) else None,
        'ano':        2024
    }
    if rec.get('operacao') and rec.get('mes'):
        meta2024.append(rec)
all_data['meta2024'] = meta2024

# =============================================================
# ABS GERAL (new source)
# =============================================================
abs_geral_data = []
try:
    if 'ABS GERAL' in wb.sheetnames:
        ws_abs = wb['ABS GERAL']
        for row in ws_abs.iter_rows(min_row=2, max_row=ws_abs.max_row, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            abs_geral_data.append({
                'ano': row[0],
                'mes': normalize_mes(row[1]),
                'abs': row[2] if isinstance(row[2], (int, float)) else 0
            })
except Exception as e:
    print(f"Warning extracting ABS GERAL: {e}")
all_data['abs_geral_timeline'] = abs_geral_data

# =============================================================
# META CNU
# =============================================================
ws = wb['META CNU']
meta_cnu = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None:
        continue
    rec = {
        'mes':        normalize_mes(row[0]),
        'operacao':   str(row[1]).strip() if row[1] else None,
        'arrecadado': row[2] if isinstance(row[2], (int, float)) else None,
        'meta':       row[3] if isinstance(row[3], (int, float)) else None,
        'alcance':    row[4] if isinstance(row[4], (int, float)) else None,
        'ano':        row[5] if isinstance(row[5], (int, float)) else 2025
    }
    if rec.get('operacao'):
        meta_cnu.append(rec)
all_data['meta_cnu'] = meta_cnu

# =============================================================
# QUARTIL
# =============================================================
ws = wb['QUARTIL']
quartil = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    vals = [serialize(v) for v in row]
    if any(v is not None for v in vals):
        quartil.append(vals)
all_data['quartil_raw'] = quartil

# =============================================================
# CAPACITY
# =============================================================
try:
    if 'CAPACITY' in wb.sheetnames:
        all_data['capacity_raw'] = [list(row) for row in wb['CAPACITY'].iter_rows(min_row=4, max_row=10, values_only=True)]
except Exception as e:
    print(f"Warning extracting CAPACITY: {e}")


# =============================================================
# CUSTOS
# =============================================================
ws = wb['CUSTOS']
custos = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    vals = [serialize(v) for v in row]
    if any(v is not None for v in vals):
        custos.append(vals)
all_data['custos'] = custos

# =============================================================
# ALARES
# =============================================================
try:
    ws_alares = wb['ALARES']
    alares_list = []
    
    # Get headers to map column indices
    headers = [str(cell.value).upper().strip() if cell.value else "" for cell in ws_alares[1]]
    idx_ano = headers.index('ANO') if 'ANO' in headers else 0
    idx_mes = headers.index('MÊS') if 'MÊS' in headers else (headers.index('MES') if 'MES' in headers else 1)
    idx_var = headers.index('VARIAVEL') if 'VARIAVEL' in headers else 2
    idx_dig = headers.index('DIGITAL') if 'DIGITAL' in headers else 3
    idx_paf = headers.index('PA FIXA') if 'PA FIXA' in headers else 4
    idx_vaf = headers.index('VARIAVEL + FIXA') if 'VARIAVEL + FIXA' in headers else -1
    idx_tot = headers.index('VALOR FINAL') if 'VALOR FINAL' in headers else 5

    for row in ws_alares.iter_rows(min_row=2, max_row=ws_alares.max_row, values_only=True):
        if not row or row[idx_ano] is None or row[idx_mes] is None:
            continue
        
        val_var = row[idx_var] if isinstance(row[idx_var], (int, float)) else 0
        val_dig = row[idx_dig] if isinstance(row[idx_dig], (int, float)) else 0
        val_paf = row[idx_paf] if isinstance(row[idx_paf], (int, float)) else 0
        val_vaf = row[idx_vaf] if idx_vaf >= 0 and isinstance(row[idx_vaf], (int, float)) else (val_var + val_paf)
        val_tot = row[idx_tot] if isinstance(row[idx_tot], (int, float)) else (val_vaf + val_dig)
        
        alares_list.append({
            'ano':      row[idx_ano],
            'mes':      normalize_mes(row[idx_mes]),
            'variavel': val_var,
            'digital':  val_dig,
            'pa_fixa':  val_paf,
            'var_fixa': val_vaf,
            'total':    val_tot
        })
    all_data['alares'] = alares_list
    print(f"  ALARES:              {len(alares_list)}")
except Exception as e:
    print(f"  Warning: Could not process ALARES sheet: {e}")
    all_data['alares'] = []

# =============================================================
# AGORACRED
# =============================================================
try:
    if 'FATURAMENTO AGORACRED- PLANO' in wb.sheetnames:
        ws_agoracred = wb['FATURAMENTO AGORACRED- PLANO']
        agoracred_list = []
        last_fat_plano = 0
        last_fat_produzido = 0
        last_mes_pg = None
        
        for row in ws_agoracred.iter_rows(min_row=2, max_row=ws_agoracred.max_row, values_only=True):
            if row[0] is None and row[1] is None:
                continue
            
            periodo = serialize(row[0])
            mes_op = serialize(row[1])
            mes_pg_nota = serialize(row[2])
            fat_quinzena = row[3] if isinstance(row[3], (int, float)) else 0
            imposto_nota = row[4] if isinstance(row[4], (int, float)) else 0
            numero_nf = serialize(row[5])
            
            current_fat_plano = row[6] if isinstance(row[6], (int, float)) else None
            current_fat_produzido = row[7] if isinstance(row[7], (int, float)) else None
            
            fat_plano = current_fat_plano if current_fat_plano is not None else 0
            fat_produzido = current_fat_produzido if current_fat_produzido is not None else 0
            
            # Simple month/year extraction for filters if possible
            mes_norm = normalize_mes(row[2]) if row[2] else normalize_mes(row[0])
            
            ano = ''
            if isinstance(row[2], datetime):
                ano = str(row[2].year)
            elif isinstance(row[2], str):
                match = re.search(r'\b(202\d)\b', row[2])
                if match: ano = match.group(1)
            
            if not ano and isinstance(row[0], str):
                match = re.search(r'\b(202\d)\b', row[0])
                if match: ano = match.group(1)
            
            agoracred_list.append({
                'periodo': periodo,
                'mes_op': mes_op,
                'mes_pg_nota': mes_pg_nota,
                'mes_norm': mes_norm,
                'ano': ano,
                'fat_quinzena': fat_quinzena,
                'imposto_nota': imposto_nota,
                'numero_nf': numero_nf,
                'fat_plano': fat_plano,
                'fat_produzido': fat_produzido
            })
        all_data['agoracred'] = agoracred_list
        print(f"  AGORACRED:           {len(agoracred_list)}")
except Exception as e:
    print(f"  Warning: Could not process AGORACRED sheet: {e}")
    all_data['agoracred'] = []



# =============================================================
# EXTRAÇÃO DE DISPERSÃO E QUARTIL (META GERAL 2026.xlsx)
# =============================================================
print("  Extraindo dados de Dispersão da planilha META GERAL 2026...")
try:
    meta_path = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\META GERAL 2026.xlsx'
    wb_meta = openpyxl.load_workbook(meta_path, data_only=True)
    
    # Filtrar abas que começam com "METAS " e têm "2026"
    meta_sheets = [s for s in wb_meta.sheetnames if s.startswith('METAS') and '2026' in s and 'Backup' not in s]
    
    mes_name_map = {
        'JANEIRO': 'janeiro', 'FEVEREIRO': 'fevereiro', 'MARCO': 'marco', 'MARÇO': 'marco',
        'ABRIL': 'abril', 'MAIO': 'maio', 'JUNHO': 'junho', 'JULHO': 'julho',
        'AGOSTO': 'agosto', 'SETEMBRO': 'setembro', 'OUTUBRO': 'outubro',
        'NOVEMBRO': 'novembro', 'DEZEMBRO': 'dezembro'
    }
    
    dispersao_data = {}
    
    for sh in meta_sheets:
        # Tentar descobrir o mês a partir do nome da aba
        sh_upper = strip_acc(sh).replace('METAS', '').replace('2026', '').strip()
        mes_key = mes_name_map.get(sh_upper, None)
        if not mes_key:
            for k, v in mes_name_map.items():
                if k in strip_acc(sh):
                    mes_key = v
                    break
        if not mes_key:
            continue
            
        ws_meta = wb_meta[sh]
        
        # O cabeçalho geralmente está na linha 4
        row_iter = ws_meta.iter_rows(min_row=4, max_row=ws_meta.max_row, values_only=True)
        try:
            first_row = next(row_iter)
        except StopIteration:
            continue
        
        headers = [strip_acc(str(h)) if h else '' for h in first_row]
        
        # Encontrar as colunas
        idx_agente = -1
        idx_op = -1
        idx_ho = -1
        idx_prom = -1
        
        for i, h in enumerate(headers):
            if 'AGENTE' in h: idx_agente = i
            elif 'OPERACAO' in h: idx_op = i
            elif 'H O' in h or h == 'HO': idx_ho = i
            elif 'PROMESSA' in h: idx_prom = i
            
        if idx_agente == -1 or idx_op == -1:
            continue
            
        ops_dict = {}
        empty_rows = 0
        for row in row_iter:
            if not any(row):
                empty_rows += 1
                if empty_rows > 50:
                    break
                continue
            empty_rows = 0
            
            agente = str(row[idx_agente]).strip() if idx_agente < len(row) and row[idx_agente] else None
            op = str(row[idx_op]).strip() if idx_op < len(row) and row[idx_op] else None
            
            if not agente or not op or agente.lower() == 'nan':
                continue
            
            ho = 0
            if idx_ho != -1 and idx_ho < len(row) and row[idx_ho] is not None:
                try:
                    ho = float(row[idx_ho])
                except:
                    ho = 0
            
            prom = 0
            if idx_prom != -1 and idx_prom < len(row) and row[idx_prom] is not None:
                try:
                    prom = float(row[idx_prom])
                except:
                    prom = 0
            
            # If HO is zero but Promessa is available, map Promessa to HO for display purposes
            if ho == 0 and prom > 0:
                ho = prom
            
            if op not in ops_dict:
                ops_dict[op] = []
            
            ops_dict[op].append({
                'agente': agente,
                'ho': ho,
                'promessas': prom
            })
            
        # Agora calcular quartil e dispersão por operação
        mes_result = {}
        
        for op, ops_list in ops_dict.items():
            # HO
            validos_ho = [x for x in ops_list if x['ho'] > 0]
            validos_ho.sort(key=lambda x: x['ho'], reverse=True)
            max_ho = max([x['ho'] for x in validos_ho], default=0)
            n_ho = len(validos_ho)
            
            for rank, r in enumerate(validos_ho):
                pct = rank / max(1, n_ho - 1)
                if n_ho <= 1: q = 1
                elif pct <= 0.25: q = 1
                elif pct <= 0.50: q = 2
                elif pct <= 0.75: q = 3
                else: q = 4
                
                r['_q_ho'] = f"{q}º Quartil"
                r['_disp_ho'] = round((r['ho'] / max_ho * 100), 1) if max_ho > 0 else 0
                
            # Promessas
            validos_prom = [x for x in ops_list if x['promessas'] > 0]
            validos_prom.sort(key=lambda x: x['promessas'], reverse=True)
            max_prom = max([x['promessas'] for x in validos_prom], default=0)
            n_prom = len(validos_prom)
            
            for rank, r in enumerate(validos_prom):
                pct = rank / max(1, n_prom - 1)
                if n_prom <= 1: q = 1
                elif pct <= 0.25: q = 1
                elif pct <= 0.50: q = 2
                elif pct <= 0.75: q = 3
                else: q = 4
                
                r['_q_prom'] = f"{q}º Quartil"
                r['_disp_prom'] = round((r['promessas'] / max_prom * 100), 1) if max_prom > 0 else 0
                
            # Keep ALL ops, not just valid ones
            mes_result[op] = ops_list
            
        dispersao_data[mes_key] = mes_result
        print(f"    {sh}: {len(ops_dict)} operações processadas.")
        
    all_data['dispersao_data'] = dispersao_data
    print("  Dispersão OK!")
    
except Exception as e:
    print(f"  Warning: Falha ao extrair dispersao de META GERAL 2026: {e}")
    import traceback
    traceback.print_exc()


# =============================================================
# Write data.js
# =============================================================
out_path = r'c:\Users\sup.luciana\Desktop\AntiGravity\PAINEL GERAL\data_v2.js'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('const DASHBOARD_DATA = ')
    json.dump(all_data, f, ensure_ascii=False, default=str, indent=None)
    f.write(';')

print(f"\nData exported successfully!")
print(f"  Operadores:          {len(operators)}")
print(f"  Producao Operacoes:  {len(prod_ops)}")
print(f"  Fechamentos 2026:    {len(fechamentos2026)}")
print(f"  Meta 2025:           {len(meta2025)}")
print(f"  Meta 2024:           {len(meta2024)}")
print(f"  Meta CNU:            {len(meta_cnu)}")
print(f"  ALARES:              {len(all_data.get('alares', []))}")
