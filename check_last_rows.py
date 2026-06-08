import openpyxl
XLSX = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['PRODUÇÃO OPERADORES']
max_r = ws.max_row
print(f"Max row: {max_r}")
for i in range(max_r - 20, max_r + 1):
    row = [c.value for c in ws[i]]
    print(f"Row {i}: {row[24]} | {row[0]}")
