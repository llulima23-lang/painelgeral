import openpyxl
from datetime import datetime

XLSX = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['PRODUÇÃO OPERADORES']

header = [str(c.value) for c in ws[1]]
abs_idx = -1
mes_idx = -1
for i, h in enumerate(header):
    if h.upper() == 'ABS': abs_idx = i
    if h.upper() == 'MÊS': mes_idx = i

print(f"ABS Index: {abs_idx}, MÊS Index: {mes_idx}")

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    mes_val = row[mes_idx]
    if isinstance(mes_val, datetime) and mes_val.year == 2026 and mes_val.month == 4:
        print(f"Row {i+2}: Agente={row[0]}, ABS={row[abs_idx]}")
    elif isinstance(mes_val, str) and 'ABRIL' in mes_val.upper() and '2026' in mes_val:
        print(f"Row {i+2}: Agente={row[0]}, ABS={row[abs_idx]}")
