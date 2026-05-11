import openpyxl
XLSX = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['PRODUÇÃO OPERADORES']

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    if row[0] and 'Mayane Lopes' in str(row[0]):
        print(f"Row {i+2}: Agente={row[0]}, col20(Pausa)={row[20]}, col21(Alcance Pausa)={row[21]}, col22(Pausa 100%)={row[22]}")
