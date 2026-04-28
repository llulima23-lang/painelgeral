import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx', data_only=True)
print('=== SHEET NAMES ===')
print(json.dumps(wb.sheetnames, ensure_ascii=False))

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== SHEET: {sheet_name} ===')
    print(f'Dimensions: {ws.dimensions}')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(35, ws.max_row), values_only=False), 1):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                vals.append(f'[{cell.column_letter}{cell.row}]={repr(v)}')
        if vals:
            sep = '  |  '
            print(f'  Row {row_idx}: {sep.join(vals)}')
