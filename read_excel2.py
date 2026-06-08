import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx', data_only=True)

# Get all sheet names
print('=== ALL SHEET NAMES ===')
for i, name in enumerate(wb.sheetnames):
    print(f'  {i+1}. {name}')

# For each sheet, print ALL data
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n\n{"="*60}')
    print(f'SHEET: {sheet_name}')
    print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
    print(f'{"="*60}')
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
        vals = []
        for cell in row:
            v = cell.value
            if v is not None:
                col = cell.column_letter
                vals.append(f'{col}{cell.row}={repr(v)}')
        if vals:
            sep = ' | '
            print(sep.join(vals))
