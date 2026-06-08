import openpyxl
import re
from datetime import datetime

XLSX = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx'

wb = openpyxl.load_workbook(XLSX, data_only=True)

def check_sheet(sheet_name, month_col_idx):
    print(f"\n--- Checking sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    found_any = False
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        if i == 0: # Header
            print(f"Header: {row}")
            continue
        
        val = row[month_col_idx]
        if val:
            # Check if it looks like April 2026
            s_val = str(val).upper()
            if 'ABRIL' in s_val and '2026' in s_val:
                print(f"Found row {i+1}: {row}")
                found_any = True
            elif isinstance(val, datetime) and val.month == 4 and val.year == 2026:
                print(f"Found row {i+1} (datetime): {row}")
                found_any = True
    
    if not found_any:
        print("No April 2026 rows found with current logic.")

check_sheet('PRODUÇÃO OPERADORES', 24) # mes is header[24]
check_sheet('PRODUÇÃO OPERAÇÕES', 13)  # mes is col 13
