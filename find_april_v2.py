import openpyxl
import re
from datetime import datetime

XLSX = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx'
wb = openpyxl.load_workbook(XLSX, data_only=True)

def check_sheet(sheet_name):
    print(f"\n--- Checking sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    header = [str(c.value) for c in ws[1]]
    print(f"Header: {header}")
    
    # Try to find 'mes' or 'mês' in header
    mes_idx = -1
    for i, h in enumerate(header):
        if h.lower() in ['mes', 'mês']:
            mes_idx = i
            break
    
    if mes_idx == -1:
        print("Could not find month column!")
        return

    count = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
        val = row[mes_idx]
        if val:
            if isinstance(val, datetime):
                if val.year == 2026 and val.month == 4:
                    if count < 5:
                        print(f"Row {i+2}: {row}")
                    count += 1
            elif isinstance(val, str):
                if 'ABRIL' in val.upper() and '2026' in val:
                    if count < 5:
                        print(f"Row {i+2}: {row}")
                    count += 1
    
    print(f"Total rows for April 2026: {count}")

check_sheet('PRODUÇÃO OPERADORES')
check_sheet('PRODUÇÃO OPERAÇÕES')
