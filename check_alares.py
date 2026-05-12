import openpyxl
XLSX = r'C:\Users\sup.luciana\Meu Drive\MF\MF\Indicadores de Cobrança\PRODUÇÃO EQUIPES.xlsx'
wb = openpyxl.load_workbook(XLSX, read_only=True)
print(f"Sheet names: {wb.sheetnames}")
if 'ALARES' in wb.sheetnames:
    ws = wb['ALARES']
    print("\nFirst 5 rows of ALARES:")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        print(row)
else:
    print("\nALARES sheet NOT found!")
